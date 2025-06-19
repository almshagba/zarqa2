from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from database import db
from models import User, UserLog
from routes.auth_routes import admin_required
from utils import get_user_logs, get_logs_count, get_logs_by_module, get_failed_logs, log_user_activity
from datetime import datetime, timedelta
from sqlalchemy import func, desc

admin = Blueprint('admin', __name__, url_prefix='/admin')

@admin.route('/logs')
@admin_required
def view_logs():
    """صفحة عرض سجلات المستخدمين للمدير"""
    
    # الحصول على معاملات التصفية من الطلب
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    user_id = request.args.get('user_id', type=int)
    module = request.args.get('module')
    action = request.args.get('action')
    status = request.args.get('status')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # بناء الاستعلام
    query = UserLog.query.join(User)
    
    # تطبيق المرشحات
    if user_id:
        query = query.filter(UserLog.user_id == user_id)
    
    if module:
        query = query.filter(UserLog.module == module)
    
    if action:
        query = query.filter(UserLog.action.contains(action))
    
    if status:
        query = query.filter(UserLog.status == status)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(UserLog.created_at >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(UserLog.created_at < date_to_obj)
        except ValueError:
            pass
    
    # ترتيب النتائج
    query = query.order_by(desc(UserLog.created_at))
    
    # تطبيق التصفح
    logs = query.paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    # الحصول على قائمة المستخدمين للفلترة
    users = User.query.order_by(User.username).all()
    
    # الحصول على قائمة الوحدات المتاحة
    modules = db.session.query(UserLog.module).distinct().all()
    modules = [m[0] for m in modules if m[0]]
    
    # الحصول على قائمة الأنشطة المتاحة
    actions = db.session.query(UserLog.action).distinct().all()
    actions = [a[0] for a in actions if a[0]]
    
    # تسجيل نشاط عرض السجلات
    log_user_activity(
        user_id=session['user_id'],
        action='عرض سجلات المستخدمين',
        module='الإدارة',
        description='عرض صفحة سجلات أنشطة المستخدمين'
    )
    
    return render_template('admin/logs.html', 
                         logs=logs, 
                         users=users, 
                         modules=modules,
                         actions=actions,
                         current_filters={
                             'user_id': user_id,
                             'module': module,
                             'action': action,
                             'status': status,
                             'date_from': date_from,
                             'date_to': date_to,
                             'per_page': per_page
                         })

@admin.route('/logs/stats')
@admin_required
def logs_stats():
    """إحصائيات السجلات"""
    
    # إحصائيات عامة
    total_logs = UserLog.query.count()
    failed_logs = UserLog.query.filter_by(status='failed').count()
    
    # إحصائيات آخر 7 أيام
    week_ago = datetime.utcnow() - timedelta(days=7)
    recent_logs = UserLog.query.filter(UserLog.created_at >= week_ago).count()
    
    # إحصائيات حسب المستخدم (أكثر 10 مستخدمين نشاطاً)
    user_stats = db.session.query(
        User.username,
        User.full_name,
        func.count(UserLog.id).label('log_count')
    ).join(UserLog).group_by(User.id).order_by(desc('log_count')).limit(10).all()
    
    # إحصائيات حسب الوحدة
    module_stats = db.session.query(
        UserLog.module,
        func.count(UserLog.id).label('log_count')
    ).group_by(UserLog.module).order_by(desc('log_count')).all()
    
    # إحصائيات حسب النشاط
    action_stats = db.session.query(
        UserLog.action,
        func.count(UserLog.id).label('log_count')
    ).group_by(UserLog.action).order_by(desc('log_count')).limit(10).all()
    
    # تسجيل نشاط عرض الإحصائيات
    log_user_activity(
        user_id=session['user_id'],
        action='عرض إحصائيات السجلات',
        module='الإدارة',
        description='عرض صفحة إحصائيات سجلات المستخدمين'
    )
    
    return render_template('admin/logs_stats.html',
                         total_logs=total_logs,
                         failed_logs=failed_logs,
                         recent_logs=recent_logs,
                         user_stats=user_stats,
                         module_stats=module_stats,
                         action_stats=action_stats)

@admin.route('/logs/clear', methods=['POST'])
@admin_required
def clear_logs():
    """مسح السجلات القديمة"""
    
    days = request.form.get('days', 30, type=int)
    
    if days < 1:
        flash('يجب أن يكون عدد الأيام أكبر من صفر', 'danger')
        return redirect(url_for('admin.view_logs'))
    
    # حساب التاريخ المحدد
    cutoff_date = datetime.utcnow() - timedelta(days=days)
    
    # حذف السجلات القديمة
    deleted_count = UserLog.query.filter(UserLog.created_at < cutoff_date).delete()
    db.session.commit()
    
    # تسجيل نشاط مسح السجلات
    log_user_activity(
        user_id=session['user_id'],
        action='مسح السجلات القديمة',
        module='الإدارة',
        description=f'تم مسح {deleted_count} سجل أقدم من {days} يوم'
    )
    
    flash(f'تم مسح {deleted_count} سجل بنجاح', 'success')
    return redirect(url_for('admin.view_logs'))

@admin.route('/logs/export')
@admin_required
def export_logs():
    """تصدير السجلات إلى ملف Excel"""
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from flask import make_response
    import io
    
    # الحصول على المرشحات
    user_id = request.args.get('user_id', type=int)
    module = request.args.get('module')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # بناء الاستعلام
    query = UserLog.query.join(User)
    
    if user_id:
        query = query.filter(UserLog.user_id == user_id)
    if module:
        query = query.filter(UserLog.module == module)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(UserLog.created_at >= date_from_obj)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(UserLog.created_at < date_to_obj)
        except ValueError:
            pass
    
    logs = query.order_by(desc(UserLog.created_at)).all()
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "سجلات المستخدمين"
    
    # إعداد العناوين
    headers = ['التاريخ والوقت', 'المستخدم', 'النشاط', 'الوحدة', 'الوصف', 'الحالة', 'عنوان IP']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # إضافة البيانات
    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=log.user.username)
        ws.cell(row=row, column=3, value=log.action)
        ws.cell(row=row, column=4, value=log.module)
        ws.cell(row=row, column=5, value=log.description or '')
        ws.cell(row=row, column=6, value=log.status)
        ws.cell(row=row, column=7, value=log.ip_address or '')
    
    # حفظ الملف في الذاكرة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # تسجيل نشاط التصدير
    log_user_activity(
        user_id=session['user_id'],
        action='تصدير السجلات',
        module='الإدارة',
        description=f'تصدير {len(logs)} سجل إلى ملف Excel'
    )
    
    # إنشاء الاستجابة
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=user_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response