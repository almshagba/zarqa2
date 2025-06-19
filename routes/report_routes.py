from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from database import db
from models import Employee, School, Transfer, User, FormTemplate, Leave, TechnicalDeficiency
from datetime import datetime
from sqlalchemy import func, and_, or_
import pandas as pd
from io import BytesIO
from routes.auth_routes import admin_required
from constants import JOB_TITLES, DIRECTORATE_DEPARTMENTS

report = Blueprint('report', __name__, url_prefix='/reports')

@report.route('/')
def reports_index():
    """صفحة التقارير الرئيسية"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # إحصائيات عامة
    total_employees = Employee.query.filter_by(is_directorate_employee=False).count()
    schools = School.query.all()
    
    # إحصائيات حسب المدرسة
    school_stats = db.session.query(
        School.name,
        func.count(Employee.id).label('employee_count')
    ).join(Employee).filter(
        Employee.is_directorate_employee == False
    ).group_by(School.id, School.name).all()
    
    # إحصائيات حسب الوظيفة
    job_stats = db.session.query(
        Employee.job_title,
        func.count(Employee.id).label('employee_count')
    ).filter(
        Employee.is_directorate_employee == False
    ).group_by(Employee.job_title).all()
    
    # إحصائيات حسب المبحث
    subject_stats = db.session.query(
        Employee.subject,
        func.count(Employee.id).label('employee_count')
    ).filter(
        and_(Employee.is_directorate_employee == False, Employee.subject.isnot(None))
    ).group_by(Employee.subject).all()
    
    return render_template('reports/index.html',
                         total_employees=total_employees,
                         schools=schools,
                         school_stats=school_stats,
                         job_stats=job_stats,
                         subject_stats=subject_stats,
                         datetime=datetime)

@report.route('/by-school')
def report_by_school():
    """تقرير الموظفين حسب المدرسة"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # الحصول على معاملات البحث والفلترة
    search_term = request.args.get('search', '')
    gender_filter = request.args.get('gender', '')
    region_filter = request.args.get('region', '')
    
    # بناء الاستعلام الأساسي - فلترة المدارس فقط (استبعاد الأقسام الإدارية)
    from constants import DIRECTORATE_DEPARTMENTS
    query = School.query.filter(
        ~School.name.in_(DIRECTORATE_DEPARTMENTS)
    )
    
    # تطبيق البحث بالاسم
    if search_term:
        query = query.filter(School.name.contains(search_term))
    
    # تطبيق فلترة الجنس
    if gender_filter:
        query = query.filter(School.gender == gender_filter)
    
    # تطبيق فلترة المنطقة
    if region_filter:
        query = query.filter(School.region == region_filter)
    
    # ترتيب النتائج
    schools_data = query.order_by(
        School.gender.asc(),  # ترتيب حسب الجنس أولاً
        School.name.asc()     # ثم حسب الاسم
    ).all()
    
    # إضافة عدد الموظفين لكل مدرسة
    schools = []
    for school in schools_data:
        employee_count = Employee.query.filter_by(
            school_id=school.id,
            is_directorate_employee=False
        ).count()
        school.employee_count = employee_count
        schools.append(school)
    
    # الحصول على قائمة المناطق المتاحة للفلترة
    available_regions = db.session.query(School.region).filter(
        ~School.name.in_(DIRECTORATE_DEPARTMENTS),
        School.region.isnot(None)
    ).distinct().all()
    regions = [r[0] for r in available_regions if r[0]]
    
    return render_template('reports/by_school.html',
                         schools=schools,
                         regions=regions,
                         search_term=search_term,
                         gender_filter=gender_filter,
                         region_filter=region_filter)

@report.route('/by-school/<int:school_id>')
def report_by_school_details(school_id):
    """تفاصيل موظفي مدرسة معينة"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    export_format = request.args.get('export')
    
    school = School.query.get_or_404(school_id)
    employees = Employee.query.filter_by(
        school_id=school_id,
        is_directorate_employee=False
    ).order_by(Employee.name).all()
    
    if export_format == 'excel':
        return export_to_excel(employees, f'تقرير موظفي {school.name}')
    
    return render_template('reports/school_details.html',
                         school=school,
                         employees=employees)

@report.route('/by-subject')
def report_by_subject():
    """تقرير الموظفين حسب المبحث"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    subject = request.args.get('subject')
    export_format = request.args.get('export')
    
    # الحصول على جميع المباحث المتاحة
    subjects = db.session.query(Employee.subject).filter(
        and_(Employee.is_directorate_employee == False, Employee.subject.isnot(None))
    ).distinct().order_by(Employee.subject).all()
    subjects = [s[0] for s in subjects if s[0]]
    
    employees = []
    selected_subject = None
    
    if subject:
        selected_subject = subject
        employees = Employee.query.filter_by(
            subject=subject,
            is_directorate_employee=False
        ).order_by(Employee.name).all()
        
        if export_format == 'excel':
            return export_to_excel(employees, f'تقرير موظفي مبحث {subject}')
    
    return render_template('reports/by_subject.html',
                         subjects=subjects,
                         employees=employees,
                         selected_subject=selected_subject)

@report.route('/by-job')
def report_by_job():
    """تقرير الموظفين حسب الوظيفة"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    job_title = request.args.get('job_title')
    export_format = request.args.get('export')
    
    # الحصول على جميع الوظائف المتاحة
    jobs = db.session.query(Employee.job_title).filter(
        Employee.is_directorate_employee == False
    ).distinct().order_by(Employee.job_title).all()
    jobs = [j[0] for j in jobs]
    
    employees = []
    selected_job = None
    
    if job_title:
        selected_job = job_title
        employees = Employee.query.filter_by(
            job_title=job_title,
            is_directorate_employee=False
        ).order_by(Employee.name).all()
        
        if export_format == 'excel':
            return export_to_excel(employees, f'تقرير موظفي وظيفة {job_title}')
    
    return render_template('reports/by_job.html',
                         jobs=jobs,
                         employees=employees,
                         selected_job=selected_job)

@report.route('/comprehensive')
def comprehensive_report():
    """تقرير شامل لجميع الموظفين"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    export_format = request.args.get('export')
    
    employees = Employee.query.filter_by(
        is_directorate_employee=False
    ).join(School).order_by(School.name, Employee.name).all()
    
    if export_format == 'excel':
        return export_to_excel(employees, 'التقرير الشامل لموظفي المدارس')
    
    return render_template('reports/comprehensive.html', employees=employees)

@report.route('/technical-deficiency-management')
def technical_deficiency_management():
    """صفحة إدارة النقص الفني"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # الحصول على بيانات النقص التي فيها نقص أو زيادة فقط
    deficiencies = TechnicalDeficiency.query.join(School).filter(
        db.or_(
            TechnicalDeficiency.deficiency_bachelor > 0,
            TechnicalDeficiency.deficiency_diploma > 0,
            TechnicalDeficiency.surplus_bachelor > 0,
            TechnicalDeficiency.surplus_diploma > 0
        )
    ).order_by(School.name).all()
    
    # الحصول على قوائم المدارس والتخصصات والوظائف
    schools = School.query.filter(~School.name.in_(DIRECTORATE_DEPARTMENTS)).order_by(School.name).all()
    
    # الحصول على التخصصات والمباحث من الموظفين الحاليين
    specializations = set()
    subjects = set()
    
    for emp in Employee.query.filter(Employee.is_directorate_employee == False).all():
        if emp.bachelor_specialization:
            specializations.add(emp.bachelor_specialization)
        if emp.masters_specialization:
            specializations.add(emp.masters_specialization)
        if emp.phd_specialization:
            specializations.add(emp.phd_specialization)
        if emp.high_diploma_specialization:
            specializations.add(emp.high_diploma_specialization)
        if emp.subject:
            subjects.add(emp.subject)
    
    return render_template('reports/technical_deficiency_management.html',
                         deficiencies=deficiencies,
                         schools=schools,
                         specializations=sorted(specializations),
                         jobs=JOB_TITLES,
                         subjects=sorted(subjects))

@report.route('/add-technical-deficiency', methods=['POST'])
def add_technical_deficiency():
    """إضافة نقص فني جديد"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    try:
        school_id = request.form.get('school_id')
        specialization = request.form.get('specialization')
        job_title = request.form.get('job_title')
        subject = request.form.get('subject')
        deficiency_bachelor = int(request.form.get('deficiency_bachelor', 0))
        deficiency_diploma = int(request.form.get('deficiency_diploma', 0))
        surplus_bachelor = int(request.form.get('surplus_bachelor', 0))
        surplus_diploma = int(request.form.get('surplus_diploma', 0))
        notes = request.form.get('notes')
        
        # التحقق من عدم وجود سجل مشابه
        existing = TechnicalDeficiency.query.filter_by(
            school_id=school_id,
            specialization=specialization,
            job_title=job_title,
            subject=subject
        ).first()
        
        if existing:
            flash('يوجد سجل مشابه لهذه المدرسة والتخصص والوظيفة', 'error')
            return redirect(url_for('report.technical_deficiency_management'))
        
        # إنشاء سجل جديد
        deficiency = TechnicalDeficiency(
            school_id=school_id,
            specialization=specialization,
            job_title=job_title,
            subject=subject,
            required_count=0,
            current_count=0,
            deficiency_count=0,
            deficiency_bachelor=deficiency_bachelor,
            deficiency_diploma=deficiency_diploma,
            surplus_bachelor=surplus_bachelor,
            surplus_diploma=surplus_diploma,
            notes=notes
        )
        
        db.session.add(deficiency)
        db.session.commit()
        
        flash('تم إضافة النقص الفني بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في إضافة النقص الفني: {str(e)}', 'error')
    
    return redirect(url_for('report.technical_deficiency_management'))

@report.route('/edit-technical-deficiency/<int:deficiency_id>', methods=['GET'])
def edit_technical_deficiency_page(deficiency_id):
    """عرض صفحة تعديل النقص الفني"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    deficiency = TechnicalDeficiency.query.get_or_404(deficiency_id)
    
    return render_template('reports/edit_technical_deficiency.html', 
                         deficiency=deficiency,
                         job_titles=JOB_TITLES)

@report.route('/edit-technical-deficiency/<int:deficiency_id>', methods=['POST'])
def edit_technical_deficiency(deficiency_id):
    """تعديل نقص فني"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    try:
        deficiency = TechnicalDeficiency.query.get_or_404(deficiency_id)
        
        deficiency.specialization = request.form.get('specialization')
        deficiency.job_title = request.form.get('job_title')
        deficiency.subject = request.form.get('subject')
        deficiency.deficiency_bachelor = int(request.form.get('deficiency_bachelor', 0))
        deficiency.deficiency_diploma = int(request.form.get('deficiency_diploma', 0))
        deficiency.surplus_bachelor = int(request.form.get('surplus_bachelor', 0))
        deficiency.surplus_diploma = int(request.form.get('surplus_diploma', 0))
        deficiency.notes = request.form.get('notes')
        deficiency.updated_at = datetime.utcnow()
        
        db.session.commit()
        flash('تم تحديث النقص الفني بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في تحديث النقص الفني: {str(e)}', 'error')
    
    return redirect(url_for('report.technical_deficiency_management'))

@report.route('/delete-technical-deficiency/<int:deficiency_id>', methods=['POST'])
def delete_technical_deficiency(deficiency_id):
    """حذف نقص فني"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    try:
        deficiency = TechnicalDeficiency.query.get_or_404(deficiency_id)
        db.session.delete(deficiency)
        db.session.commit()
        
        flash('تم حذف النقص الفني بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف النقص الفني: {str(e)}', 'error')
    
    return redirect(url_for('report.technical_deficiency_management'))

# تعديل route الموقف الفني الحالي
@report.route('/technical-position')
def technical_position():
    """تقرير الموقف الفني - النواقص المدخلة يدوياً"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # الحصول على معاملات الفلترة
    specialization_filter = request.args.get('specialization', '')
    gender_filter = request.args.get('gender', '')
    school_filter = request.args.get('school', '')
    subject_filter = request.args.get('subject', '')
    job_filter = request.args.get('job', '')
    
    # بناء الاستعلام
    query = TechnicalDeficiency.query.join(School)
    
    if specialization_filter:
        query = query.filter(TechnicalDeficiency.specialization.contains(specialization_filter))
    if gender_filter:
        query = query.filter(School.gender == gender_filter)
    if school_filter:
        query = query.filter(TechnicalDeficiency.school_id == school_filter)
    if subject_filter:
        query = query.filter(TechnicalDeficiency.subject == subject_filter)
    if job_filter:
        query = query.filter(TechnicalDeficiency.job_title == job_filter)
    
    # الحصول على السجلات التي تحتوي على نواقص أو زوائد
    deficiencies = query.filter(
        or_(
            TechnicalDeficiency.deficiency_bachelor > 0,
            TechnicalDeficiency.deficiency_diploma > 0,
            TechnicalDeficiency.surplus_bachelor > 0,
            TechnicalDeficiency.surplus_diploma > 0
        )
    ).order_by(School.name).all()
    
    # الحصول على قوائم الفلترة
    all_specializations = set()
    all_subjects = set()
    all_jobs = set()
    
    for deficiency in TechnicalDeficiency.query.all():
        all_specializations.add(deficiency.specialization)
        if deficiency.subject:
            all_subjects.add(deficiency.subject)
        all_jobs.add(deficiency.job_title)
    
    from constants import DIRECTORATE_DEPARTMENTS
    all_schools = School.query.filter(~School.name.in_(DIRECTORATE_DEPARTMENTS)).all()
    
    return render_template('reports/technical_position.html',
                         deficiencies=deficiencies,
                         all_specializations=sorted(all_specializations),
                         all_subjects=sorted(all_subjects),
                         all_jobs=sorted(all_jobs),
                         all_schools=all_schools,
                         specialization_filter=specialization_filter,
                         gender_filter=gender_filter,
                         school_filter=school_filter,
                         subject_filter=subject_filter,
                         job_filter=job_filter)

@report.route('/technical-position/export')
def export_technical_position():
    """تصدير تقرير الموقف الفني إلى Excel"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # الحصول على معاملات الفلترة
    specialization_filter = request.args.get('specialization', '')
    gender_filter = request.args.get('gender', '')
    school_filter = request.args.get('school', '')
    subject_filter = request.args.get('subject', '')
    job_filter = request.args.get('job', '')
    
    # بناء الاستعلام (نفس منطق technical_position)
    query = TechnicalDeficiency.query.join(School)
    
    if specialization_filter:
        query = query.filter(TechnicalDeficiency.specialization.contains(specialization_filter))
    if gender_filter:
        query = query.filter(School.gender == gender_filter)
    if school_filter:
        query = query.filter(TechnicalDeficiency.school_id == school_filter)
    if subject_filter:
        query = query.filter(TechnicalDeficiency.subject == subject_filter)
    if job_filter:
        query = query.filter(TechnicalDeficiency.job_title == job_filter)
    
    # الحصول على السجلات التي تحتوي على نواقص أو زوائد
    deficiencies = query.filter(
        or_(
            TechnicalDeficiency.deficiency_bachelor > 0,
            TechnicalDeficiency.deficiency_diploma > 0,
            TechnicalDeficiency.surplus_bachelor > 0,
            TechnicalDeficiency.surplus_diploma > 0
        )
    ).order_by(TechnicalDeficiency.specialization, School.name).all()
    
    return export_technical_position_to_excel(deficiencies)

def export_technical_position_to_excel(deficiencies):
    """تصدير بيانات الموقف الفني إلى ملف Excel"""
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # إنشاء مصنف جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "الموقف الفني"
    
    # إعداد الخط العربي
    arabic_font = Font(name='Arial Unicode MS', size=12, bold=False)
    header_font = Font(name='Arial Unicode MS', size=12, bold=True)
    title_font = Font(name='Arial Unicode MS', size=14, bold=True)
    
    # إعداد المحاذاة
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # إعداد الحدود
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # إضافة العنوان الرئيسي
    current_date = datetime.now().strftime('%Y/%m/%d')
    ws.merge_cells('A1:H1')
    ws['A1'] = 'مديرية التربية والتعليم لمنطقة الزرقاء الثانية'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'قسم شؤون الموظفين'
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f'الموقف الفني - {current_date}'
    ws['A3'].font = header_font
    ws['A3'].alignment = center_alignment
    
    # ترك صف فارغ
    current_row = 5
    
    # إضافة رؤوس الأعمدة
    headers = ['المنطقة', 'المبحث', 'الجنس', 'العدد', 'المبحث', 'المدرسة', 'التقرير/الملاحظات']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    
    current_row += 1
    
    # دالة تحويل أسماء المناطق
    def get_region_display_name(region):
        """تحويل اسم المنطقة إلى الاسم المطلوب عرضه"""
        if not region:
            return ''
        
        region_lower = region.lower()
        if 'هاشمية' in region_lower or 'الهاشمية' in region_lower:
            return 'لواء الهاشمية'
        elif 'ظليل' in region_lower or 'بيرين' in region_lower:
            return 'لواء قصبة الزرقاء'
        elif 'ازرق' in region_lower or 'الازرق' in region_lower or 'حلابات' in region_lower or 'الحلابات' in region_lower:
            return 'بادية وسطى'
        else:
            return region
    
    # إضافة البيانات
    for deficiency in deficiencies:
        # تحديد نوع النقص/الزيادة
        deficiency_type = ""
        count = 0
        
        if deficiency.deficiency_bachelor > 0:
            deficiency_type = "نقص بكالوريوس"
            count = deficiency.deficiency_bachelor
        elif deficiency.deficiency_diploma > 0:
            deficiency_type = "نقص دبلوم"
            count = deficiency.deficiency_diploma
        elif deficiency.surplus_bachelor > 0:
            deficiency_type = "زيادة بكالوريوس"
            count = deficiency.surplus_bachelor
        elif deficiency.surplus_diploma > 0:
            deficiency_type = "زيادة دبلوم"
            count = deficiency.surplus_diploma
        
        # تحويل اسم المنطقة
        region_display = get_region_display_name(deficiency.school.region)
        
        # تحديد جنس المدرسة
        school_gender = deficiency.school.gender or 'إناث'
        
        # إضافة البيانات للصف
        row_data = [
            region_display,                  # المنطقة (محولة)
            deficiency.subject if deficiency.subject else '-',        # المبحث
            school_gender,                   # الجنس
            count,                           # العدد
            deficiency.subject if deficiency.subject else '-',        # المبحث (مكرر)
            deficiency.school.name,          # المدرسة
            deficiency.notes or ''           # الملاحظات
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = arabic_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row += 1
    
    # تعديل عرض الأعمدة
    column_widths = [15, 20, 10, 10, 20, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # حفظ الملف في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # إنشاء الاستجابة
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # تشفير اسم الملف
    filename = f'الموقف_الفني_{datetime.now().strftime("%Y_%m_%d")}.xlsx'
    encoded_filename = quote(filename, safe='')
    response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
    
    return response

def export_to_excel(employees, report_title):
    """تصدير البيانات إلى ملف Excel"""
    from urllib.parse import quote
    
    # إنشاء DataFrame
    data = []
    for emp in employees:
        data.append({
            'الرقم الوزاري': emp.ministry_number,
            'الاسم': emp.name,
            'الرقم الوطني': emp.civil_id,
            'الجنس': emp.gender,
            'رقم الهاتف': emp.phone_number or '',
            'الوظيفة': emp.job_title,
            'المدرسة': emp.school.name,
            'المؤهل': emp.qualification,
            'تخصص البكالوريوس': emp.bachelor_specialization or '',
            'تخصص الدبلوم العالي': emp.high_diploma_specialization or '',
            'تخصص الماجستير': emp.masters_specialization or '',
            'تخصص الدكتوراه': emp.phd_specialization or '',
            'المبحث الدراسي': emp.subject or '',
            'تاريخ التعيين': emp.appointment_date.strftime('%Y-%m-%d') if emp.appointment_date else '',
            'المنطقة': emp.school.region or ''
        })
    
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel في الذاكرة
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='التقرير', index=False)
        
        # تنسيق الورقة
        worksheet = writer.sheets['التقرير']
        
        # تعديل عرض الأعمدة
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # إنشاء الاستجابة مع تشفير صحيح للاسم
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # تشفير اسم الملف بـ UTF-8 وتحويله إلى URL encoding
    encoded_filename = quote(f'{report_title}.xlsx', safe='')
    response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'\'{encoded_filename}'
    
    return response